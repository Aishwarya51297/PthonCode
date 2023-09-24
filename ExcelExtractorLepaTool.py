import pandas as pd
import json
import openpyxl
import re
import os

class ExcelDataConverter:
    def __init__(self, file_path, required_columns):
        self.file_path = file_path
        self.required_columns = required_columns
        self.all_sheet_data = {}

    def create_nested_dict_from_excel(self):
        xls = pd.ExcelFile(self.file_path)
        sheet_names = xls.sheet_names
        for sheet_name in sheet_names:
            sheet_data = pd.read_excel(self.file_path, sheet_name=sheet_name, header=2, usecols=lambda col: col in self.required_columns)
            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            all_list_data = []
            sheet_data = sheet_data.dropna(how='all')

            for index, row in sheet_data.iterrows():
                all_list_data.insert(index, row.to_dict())

            self.all_sheet_data[str(sheet_name)] = all_list_data

        return json.dumps(self.all_sheet_data)
    
    def Add_Duplicates_Values_Sheet_in_New_sheet(self,excel_file,column_name):
        year_to_sheets = {}  # Dictionary to store year as key and sheets as values

        # Iterate through the sheets and their corresponding lists of dictionaries
        for sheet, entries in self.all_sheet_data.items():
            for entry in entries:
                year = entry.get(column_name)  # Get the 'Year' value from each entry
                if year is not None:  # Check if 'Year' is not None
                    if year not in year_to_sheets:
                        year_to_sheets[year] = set()  # Initialize an empty set for the year if not already present
                    year_to_sheets[year].add(sheet)  # Add the sheet to the set for that year

        # Filter the results to include only years that appear in multiple sheets
        year_to_sheets = {year: sheets for year, sheets in year_to_sheets.items() if len(sheets) > 1}

        print(year_to_sheets)


        # Specify the path to your Excel file
        excel_file_path = excel_file

        # Load the existing Excel file
        book = openpyxl.load_workbook(excel_file_path)
        key_with_same_Typenummer={}
        # Create a new sheet
        sheet_name_to_delete = column_name

        # Check if the sheet exists in the workbook
        if sheet_name_to_delete in book.sheetnames:
            # Get the sheet by name and remove it
            sheet_to_delete = book[sheet_name_to_delete]
            book.remove(sheet_to_delete)
        new_sheet = book.create_sheet(column_name)
        new_sheet['A1'] = column_name
        new_sheet['B1'] = 'Sheet Names'
        row_number = 2  # Start from row 2 to avoid overwriting headers
        for year, sheets in year_to_sheets.items():
            sheet_list = list(year_to_sheets[year])
            key_with_same_Typenummer[year] = ', '.join(year_to_sheets[year])
            # Add the data to the new sheet
            new_sheet[f'A{row_number}'] = year
            new_sheet[f'B{row_number}'] = key_with_same_Typenummer[year]
            row_number += 1
            print(year)
            print(sheet_list)

        # Save the modified Excel file
        book.save(excel_file_path)

    #replace 1 oder 2 with the Max number line 2
    def ReplaceString_with_Max_number(self,excel_file_path,key_name):

        for sheet_data in self.all_sheet_data.values():
            for entry in sheet_data:
                if entry[key_name] is not None:
                    numbers = re.findall(r'\d+', entry[key_name])
                    if numbers:
                        # Convert the numbers to integers
                        numbers = [int(num) for num in numbers]
                    
                        # Find the maximum number
                        entry[key_name] = max(numbers)

            # Print the updated data
        print(self.all_sheet_data)

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
        # Create a new Excel workbook
        book = openpyxl.Workbook(excel_file_path)
        # Iterate through the data dictionary and create sheets with data
        for sheet_name, sheet_data in self.all_sheet_data.items():
            # Create a new sheet
            sheet = book.create_sheet(title=sheet_name)
            
            # Add headers
            headers = list(sheet_data[0].keys()) if sheet_data else []
            sheet.append(headers)
            
            # # Add data
            for entry in sheet_data:
                row_data = [entry[header] for header in headers]
                sheet.append(row_data)

        # # Remove the default sheet created with the workbook
        # default_sheet = book['Sheet1']
        # book.remove(default_sheet)

        # Save the workbook to a file
        book.save(excel_file_path)







if __name__ == "__main__":
    excel_file = "172_Excel_Slicers.xlsx" # your base file name
    output_file="output.xlsx" #replace the filename you want share with the manager
    required_columns = ['Year', 'Type', 'Product group', 'Producer', 'Volume', 'Cost per unit', 'Price per unit', 'Revenue']

    converter = ExcelDataConverter(excel_file, required_columns)
    nested_dict = converter.create_nested_dict_from_excel()
    key_name='Type' # replace 'stukzul' here which is having '1 oder 2' values
    converter.ReplaceString_with_Max_number(output_file,key_name)
    key_name='Year' #replace Typenummer here
    converter.Add_Duplicates_Values_Sheet_in_New_sheet(output_file,key_name)



